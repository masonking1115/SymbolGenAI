"""Generate the test1 parts-selection BOM workbook.

Produces test1_bom.xlsx with one row per component needing selection. The
right-most column carries a self-contained, manufacturer-neutral prompt for
the parts-explorer agent.

Reflects the V-to-I bias topology decision (Voltai_Notes 2026-05-24) plus the
bias-polarity fix: BIAS0/BIAS1 implemented as MCP4728 V-DAC (external VREF=3.3V
for rail-to-rail output) -> dual RRIO op-amp + PMOS pass element + R_sense
high-side transconductance loop. PMOS drain sources current INTO BIASx; loop
shuts off cleanly at V_DAC=3.3V (MCP4728 EEPROM default 0xFFF).
"""

import glob
import os
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Per-MPN one-line descriptions for the per-refdes BOM tab. Generic Device:R /
# Device:C (no committed MPN) fall through to a value-only description.
LIB_DESC = {
    "Bobcat":                "Bobcat DUT — custom 40-QFN test chip (5×5 mm, 0.4 mm pitch, EP=GND)",
    "TPS7A8401A":            "LDO, 3 A, ANY-OUT programmable output, VQFN-20",
    "TPS22916CNYFPR":        "Load switch, 1-channel, 5.5 V / 2 A, WCSP",
    "24AA08-I-SN":           "EEPROM, 8 Kbit I²C, SOIC-8",
    "MCP4728":               "DAC, quad 12-bit voltage-output, I²C, MSOP-10",
    "OPA2388":               "Op-amp, dual precision RRIO, zero-drift, MSOP-8",
    "PMZ1200UPEYL":          "MOSFET, P-channel, 20 V, SOT-666",
    "2N7002":                "MOSFET, N-channel, 60 V, SOT-23",
    "ASP-134606-01":         "Connector, VITA 57.1 FMC LPC mezzanine, 160-pos SMT",
    "HRM-G-300-467B-1":      "Connector, SMA jack, 50 Ω, edge-launch",
    "TSW-102-05-G-S":        "Header, 1×2, 2.54 mm pitch, TH vertical",
    "TSW-104-05-G-S":        "Header, 1×4, 2.54 mm pitch, TH vertical",
    "Keystone-5011":         "Test point, GND test clip / loop, TH",
    # Passives (per-MPN .SchLib in Parts Library/)
    "GRM21BR71A106KA73L":    "MLCC, 10 µF, X7R, 10 V, 0805",
    "GRM155R70J105KA12D":    "MLCC, 1 µF, X7R, 6.3 V, 0402",
    "GRM155R71C104KA88D":    "MLCC, 100 nF, X7R, 16 V, 0402",
    "CRCW04020000Z0ED":      "Resistor, 0 Ω jumper, 0402",
    "CR0402-FX-1002GLF":     "Resistor, 10 kΩ, 1%, 1/16 W, 0402",
    "TNPW06035K11BEEA":      "Resistor, 5.11 kΩ, 0.1%, 25 ppm/°C, thin-film, 0603",
    "GRM155R71H103KA88D":    "MLCC, 10 nF, X7R, 50 V, 0402",
    "GRM21BR61A226ME44L":    "MLCC, 22 µF, X5R, 10 V, 0805",
    "CR0402-FX-1001GLF":     "Resistor, 1 kΩ, 1%, 1/16 W, 0402",
    "CR0402-FX-2201GLF":     "Resistor, 2.2 kΩ, 1%, 1/16 W, 0402",
}


def _description_for(lib_id: str, value: str) -> str:
    """One-line description for the per-refdes BOM tab.

    lib_id is the netlist's `Lib:<MPN>` (or `Device:R`/`Device:C` for the
    handful of generic passives that have no committed MPN yet)."""
    if not lib_id:
        return value or ""
    prefix, _, body = lib_id.partition(":")
    if prefix == "Lib" and body in LIB_DESC:
        return LIB_DESC[body]
    if lib_id == "Device:R":
        return f"Resistor, {value} (no MPN assigned — Device:R)"
    if lib_id == "Device:C":
        return f"Capacitor, {value} (no MPN assigned — Device:C)"
    # Unknown library id: surface what we know rather than dropping it.
    return f"{body or lib_id} ({value})" if value else (body or lib_id)


def _mpn_for(lib_id: str) -> str:
    """MPN extracted from the netlist lib_id. `Lib:<MPN>` → `<MPN>`;
    generic `Device:R`/`Device:C` (no committed MPN) → empty string."""
    prefix, _, body = (lib_id or "").partition(":")
    if prefix == "Lib":
        return body
    return ""  # Device:R, Device:C — no MPN assigned yet


def _collect_per_refdes(netlist_dir: Path) -> list[tuple[str, str, str, str, str]]:
    """Return [(refdes, value, mpn, description, sheet), ...] sorted across all
    netlist YAML files. One row per placed part instance."""
    rows: list[tuple[str, str, str, str, str]] = []
    for p in sorted(netlist_dir.glob("*.yaml")):
        sheet = p.stem
        data = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
        for ref, spec in (data.get("parts") or {}).items():
            if not isinstance(spec, dict):
                continue
            value = str(spec.get("value", "") or "")
            lib_id = str(spec.get("lib_id", "") or "")
            mpn = _mpn_for(lib_id)
            desc = _description_for(lib_id, value)
            rows.append((ref, value, mpn, desc, sheet))
    # Natural-ish sort: letter prefix then numeric suffix when present.
    def key(r):
        ref = r[0]
        head = "".join(c for c in ref if c.isalpha())
        tail = "".join(c for c in ref if c.isdigit())
        return (head, int(tail) if tail.isdigit() else 0, ref)
    return sorted(rows, key=key)

COLUMNS = [
    ("Item",                 6),
    ("Ref Des",              14),
    ("Qty",                  6),
    ("Category",             20),
    ("Description",          40),
    ("Value / Rating",       22),
    ("Package",              18),
    ("MPN",                  22),
    ("Datasheet URL",        30),
    ("voltai chat url",      30),
    ("Notes",                30),
    ("Parts Agent Prompt",   90),
]

PROMPT_LOAD_SWITCH = (
    "Select a single-channel load-switch IC.\n"
    "Electrical:\n"
    " - Input voltage range: 1.2 V to 3.3 V; must switch reliably across the full range.\n"
    " - Continuous output current: >=500 mA (load is a digital I/O rail feeding a custom IC, ~100-300 mA expected).\n"
    " - On-resistance: <=200 mOhm at 3.3 V Vin.\n"
    " - Quiescent current (enabled): <10 uA typical preferred.\n"
    " - Shutdown current: <1 uA typical preferred.\n"
    "Control:\n"
    " - Active-high digital enable input, compatible with 1.8 V and 3.3 V CMOS logic.\n"
    " - External 10 kOhm pull-down to GND will be present on the EN line.\n"
    " - Controlled slew rate / soft-start to limit inrush into bulk decoupling preferred.\n"
    " - Output discharge (active pull-down when disabled) preferred.\n"
    " - Reverse-current blocking preferred but not required.\n"
    "Mechanical / environmental:\n"
    " - Small SMT package (SOT-23-6, SC-70-6, WSON, or smaller).\n"
    " - Operating temperature: -40 C to +85 C minimum.\n"
    " - Lead-free, RoHS compliant.\n"
    "Context: routes FMC VADJ rail (1.2-3.3 V from the carrier) to the VDDIO pin of a custom QFN test chip via a 1x2 jumper."
)

PROMPT_EEPROM = (
    "Select a serial EEPROM.\n"
    "Memory:\n"
    " - Density: 8 Kbit (1024 x 8) exactly.\n"
    " - Endurance: >=1,000,000 write cycles per cell.\n"
    " - Data retention: >=100 years at 25 C typical.\n"
    "Interface:\n"
    " - I2C / 2-wire, standard mode (100 kHz) required, fast mode (400 kHz) preferred.\n"
    " - 7-bit slave address with hardware-selectable address pins (A0, A1, A2) preferred for bus sharing with other I2C peripherals.\n"
    " - Write-protect (WP) pin preferred.\n"
    "Supply:\n"
    " - Must operate at 3.3 V single supply. Wider tolerance window (e.g. 1.7-5.5 V) acceptable.\n"
    " - Standby current: <5 uA typical preferred.\n"
    "Package: SOIC-8, TSSOP-8, MSOP-8, or DFN-8.\n"
    "Temperature: -40 C to +85 C minimum.\n"
    "Context: stores VITA 57.1 FMC mezzanine board identification (IPMI FRU data). Must be conventionally usable in that role on an FMC LPC daughtercard."
)

PROMPT_BIAS_VOLTAGE_DAC = (
    "Select an I2C-controlled voltage-output DAC to drive a high-side PMOS transconductance bias loop.\n"
    "Topology context:\n"
    " - The DAC drives the non-inverting input of an external precision op-amp.\n"
    " - The op-amp output drives the gate of a PMOS pass transistor.\n"
    " - The PMOS source connects to 3.3V through a sense resistor R_sense; the source node also feeds the op-amp inverting input.\n"
    " - The PMOS drain delivers regulated current INTO the test-chip bias pin via a 1x2 jumper.\n"
    " - I_load = (3.3 V - V_DAC) / R_sense.\n"
    "Requirements:\n"
    " - Resolution: >=12-bit preferred (10-bit minimum) for ~1 uA resolution over 0-640 uA full scale.\n"
    " - Output swing: rail-to-rail at 3.3 V supply; MUST be able to reach the positive rail (~3.3 V) to fully shut off the PMOS and reach 0 uA at the low end.\n"
    " - External voltage reference input required (or selectable) so V_REF = 3.3 V can be used; do not rely on a 2.048 V internal reference.\n"
    " - Default state at power-up: output at the positive rail (PMOS off). Must be programmable for default-off behavior (EEPROM default code = 0xFFF or equivalent).\n"
    " - Settling time: <1 ms acceptable; control loop sets bandwidth.\n"
    "Interface:\n"
    " - I2C, standard or fast mode.\n"
    " - Hardware-selectable address pins preferred.\n"
    "Supply: 3.3 V single supply.\n"
    "Channels: 2 channels needed total (BIAS0 and BIAS1). A 2-channel or higher device acceptable.\n"
    "Package: any small SMT.\n"
    "Temperature: -40 C to +85 C minimum."
)

PROMPT_OP_AMP = (
    "Select a precision op-amp used in a high-side transconductance current-source loop.\n"
    "Topology: V_DAC drives the non-inverting input. A PMOS pass transistor has its source connected to 3.3 V through a sense resistor R_sense; that source node also feeds the op-amp inverting input. The op-amp output drives the PMOS gate. The PMOS drain delivers the regulated output current INTO the test-chip bias pin.\n"
    "Requirements:\n"
    " - Single supply operation at 3.3 V.\n"
    " - Rail-to-rail input AND rail-to-rail output, both required (output must drive PMOS gate close to either rail to turn it fully off and fully on).\n"
    " - Input offset voltage (Vos): <=0.5 mV maximum. Critical for accuracy at the low end of bias current (~1-10 uA).\n"
    " - Input bias current: <=1 nA typical (CMOS / FET input stage).\n"
    " - Input common-mode range must include both rails (sees V_sense which sits near 3.3 V at low bias currents).\n"
    " - Gain-bandwidth product: >=1 MHz.\n"
    " - Unity-gain stable; capable of driving a MOSFET gate (tens to hundreds of pF) without oscillation.\n"
    " - Quiescent current: <500 uA per channel preferred.\n"
    "Channels: 2 channels needed total (one per bias source). Dual-channel device preferred.\n"
    "Package: any small SMT (SOT-23-5, SC-70-5, MSOP-8, DFN-8).\n"
    "Temperature: -40 C to +85 C minimum."
)

PROMPT_PMOS = (
    "Select a small-signal P-channel MOSFET for a high-side linear current-source pass element.\n"
    "Topology: source connected to 3.3 V through a small sense resistor; gate driven by an op-amp output; drain delivers regulated current INTO the test-chip bias pin (drain sits anywhere from ~0 V to ~1 V depending on chip load).\n"
    "Requirements:\n"
    " - Vds breakdown (Vds_max): >=6 V.\n"
    " - |Vgs(th)|: logic-level threshold; |Vgs(th)| <= 2 V so the device fully enhances within the 3.3 V gate-drive range.\n"
    " - Continuous drain current rating: >=10 mA (operates at <=1 mA so any small-signal device is fine).\n"
    " - Rds(on): not critical (linear-region operation, op-amp closes the loop).\n"
    " - Total gate charge / input capacitance: low, <=200 pF preferred so the op-amp can drive the gate without instability.\n"
    " - Low gate leakage and low drain leakage in the off state (output must reach <<1 uA when commanded off).\n"
    " - Operating temperature: -40 C to +85 C minimum.\n"
    "Quantity: 2 units required.\n"
    "Package: SOT-23, SOT-323, SOT-666, SC-70, or smaller. Single-transistor part preferred."
)

PROMPT_RSENSE = (
    "Select a precision thin-film chip resistor used as a V-to-I sense element in a high-side PMOS transconductance bias loop.\n"
    "Requirements:\n"
    " - Value: 5.11 kOhm exact (E96 series).\n"
    " - Tolerance: 0.1% maximum.\n"
    " - Temperature coefficient: <=25 ppm/C.\n"
    " - Construction: thin-film (not thick-film).\n"
    " - Power rating: >=1/10 W (typical 0603); max dissipation = 3.3V * 646uA = 2.1 mW.\n"
    " - Package: 0603 preferred; 0402 acceptable if 0.1% / 25 ppm available.\n"
    " - Operating temperature: -40 C to +85 C minimum.\n"
    "Context: programs full-scale bias current at I_FS = V_supply / R_sense = 3.3 V / 5.11 kOhm ~= 646 uA over a 0-3.3V V_DAC range."
)

PROMPT_SMA = (
    "Select a board-mount SMA connector (RF coaxial, 50 ohm).\n"
    "Mechanical:\n"
    " - Mount style: edge-launch end-launch OR through-hole vertical PCB mount, both acceptable.\n"
    " - Gender: female receptacle (jack) - mates with standard SMA male plug cables.\n"
    " - Compatible with 1.6 mm to 2.4 mm PCB thickness.\n"
    " - Body material: stainless steel or brass with nickel or stainless plating.\n"
    " - Center contact: gold plated.\n"
    " - Mating cycles: >=500.\n"
    "Electrical:\n"
    " - Characteristic impedance: 50 ohm.\n"
    " - Frequency range: DC to >=6 GHz.\n"
    " - VSWR: <=1.3 to 6 GHz preferred.\n"
    " - Insertion loss: <=0.3 dB at 6 GHz.\n"
    "Quantity needed: 7 on this design.\n"
    "Temperature: -40 C to +85 C minimum."
)

PROMPT_FMC_CONNECTOR = (
    "Select the mezzanine-side connector for a VITA 57.1 FMC LPC daughtercard.\n"
    "Standard:\n"
    " - VITA 57.1 LPC (Low Pin Count), 160 positions in a 10 column x 4 row grid (rows C, D, G, H populated; rows A, B, E, F unpopulated for LPC).\n"
    " - Pitch: 1.27 mm.\n"
    " - Polarization / keying per VITA 57.1 LPC.\n"
    "Role:\n"
    " - This board is the MEZZANINE (the daughtercard). Select the mating half that goes on the mezzanine PCB (not the carrier socket).\n"
    " - Must mate with the FMC LPC socket on a Digilent Genesys 2 carrier board.\n"
    "Mount:\n"
    " - Surface-mount on the bottom side of the mezzanine PCB.\n"
    " - Standard FMC stack height (8.5 mm or 10 mm). Confirm Genesys 2 compatibility - Genesys 2 typically expects 8.5 mm.\n"
    "Electrical:\n"
    " - Signal-pair speed rating: >=10 Gbps per differential pair (per VITA 57.1 LPC).\n"
    " - Power pin current rating per VITA 57.1 (>=2 A per power contact typical).\n"
    "Temperature: -40 C to +85 C minimum.\n"
    "Quantity: 1."
)

PROMPT_HEADER_1X4 = (
    "Select a 1x4 single-row male pin header.\n"
    " - Pitch: 2.54 mm (100 mil).\n"
    " - Positions: 4.\n"
    " - Rows: 1.\n"
    " - Mount: through-hole, vertical (pins straight up off the board).\n"
    " - Gender: male pins (mates with standard 0.1-inch female jumper-wire sockets).\n"
    " - Pin material: brass or phosphor bronze with tin or gold plating.\n"
    " - Current rating: >=1 A per pin.\n"
    " - Temperature: -40 C to +85 C minimum.\n"
    "Context: breakout for 4 GPIO test signals from a custom QFN test chip."
)

PROMPT_JUMPER_1X2 = (
    "Select a 1x2 (2-position, 1-row) male pin header AND its mating shunt cap.\n"
    " - Pitch: 2.54 mm (100 mil).\n"
    " - Mount: through-hole, vertical.\n"
    " - Pin material: brass or phosphor bronze with tin or gold plating.\n"
    " - Current rating: >=1 A.\n"
    " - Mating shunt: open-top or closed-top, gold or tin plated, line up with selected header.\n"
    " - Temperature: -40 C to +85 C minimum.\n"
    "Context: used as a manual disconnect on multiple board nets (LDO output legs, bias source outputs, VDDIO supply). Will be installed by default; the shunt is removed during fault isolation."
)

PROMPT_DECOUPLE_CAP = (
    "Select MLCC bypass capacitors for digital power-rail decoupling.\n"
    "Three-value bank per Bobcat power pin (10 uF + 1 uF + 100 nF):\n"
    " - Dielectric: X7R or X5R only. Y5V and Z5U are NOT acceptable due to large CV droop.\n"
    " - Voltage rating: >=6.3 V for all (rails are 0.6 V to 3.3 V).\n"
    " - Tolerance: +/-10% acceptable; +/-20% acceptable for the bulk 10 uF.\n"
    " - Package size: 0402 preferred for 100 nF and 1 uF; 0603 or 0805 acceptable for 10 uF.\n"
    " - ESR / ESL: standard general-purpose MLCC; no special low-ESR or controlled-ESR variant required.\n"
    " - Operating temperature: -40 C to +85 C minimum.\n"
    "Quantity: see Qty column; one bank per Bobcat supply pin (VDDD, VDDIO, VDDA1, VDDA2)."
)

PROMPT_ZERO_OHM = (
    "Select 0 ohm chip jumper resistors.\n"
    " - Value: 0 ohm.\n"
    " - Package: 0402 preferred (or 0603 if 0402 introduces sourcing risk; pick one size and use everywhere).\n"
    " - Current rating: >=1 A.\n"
    " - Operating temperature: -40 C to +125 C.\n"
    "Used in three roles on this board: series jumpers on FMC LA digital signals, supply series elements on analog rails, and optional jumpers between FMC and SMA outputs."
)

PROMPT_10K_PULL = (
    "Select 10 kOhm chip pull-up / pull-down resistors.\n"
    " - Value: 10 kOhm.\n"
    " - Tolerance: +/-5% or tighter (+/-1% acceptable).\n"
    " - Power rating: >=1/16 W (typical for 0402).\n"
    " - Package: 0402 preferred (or 0603 if 0402 introduces sourcing risk; uniform with the 0 ohm choice).\n"
    " - Operating temperature: -40 C to +125 C.\n"
    "Used as default-state biasing on digital control lines and as discharge paths."
)

PROMPT_TEST_POINT = (
    "Select PCB-mount GND test points / test loops.\n"
    " - Mount: through-hole.\n"
    " - Compatible with standard 0.1-inch oscilloscope probe ground spring clips and standard mini-grabber clip leads.\n"
    " - Loop style or turret pin style - both acceptable.\n"
    " - Color coding: black (GND) preferred where available, otherwise unmarked is fine.\n"
    " - Current rating: not critical (probing only).\n"
    " - Quantity: 4 distributed across the board for easy probe access."
)

ROWS = [
    # (Ref, Qty, Category, Description, Value, Package, MPN, prompt)
    ("U0",       1,  "DUT (test chip)",
        "Bobcat - custom mixed-signal test chip (Unconventional AI). Device under test for this carrier board.",
        "40-QFN, 5x5 mm, 0.4 mm pitch, exposed GND pad (pin 41)", "QFN-40 5x5 mm",
        "Bobcat",
        "N/A - DUT is customer-specified, not selected via parts agent. Blackbox component; no part datasheet available yet. Pin/electrical spec source = '[External] Bobcat Board Design.pdf' in project root."),
    ("U1",       1,  "Linear regulator",
        "High-accuracy 3 A adjustable LDO with ANY-OUT pin-programmable output",
        "ANY-OUT 0.5-2.075 V", "VQFN-20 3.5x3.5 mm",
        "TPS7A8401A",
        "Already selected. Confirm orderable variant (RGW vs RGR) and stock."),
    ("U2",       1,  "Load switch",
        "Single-channel load switch for VADJ -> Bobcat VDDIO",
        "Iout >= 500 mA, EN active-high", "SOT-23-6 or smaller",
        "", PROMPT_LOAD_SWITCH),
    ("U3",       1,  "EEPROM",
        "8 Kbit I2C EEPROM for FMC FRU data",
        "8 Kbit, I2C, 3.3 V", "SOIC-8 or smaller",
        "", PROMPT_EEPROM),
    ("U4",       1,  "Voltage DAC (bias)",
        "Quad 12-bit I2C voltage-output DAC; external VREF=3.3V for rail-to-rail output, drives PMOS V-to-I loop",
        "12-bit, 0-3.3 V (ext VREF), I2C", "MSOP-10 or smaller",
        "", PROMPT_BIAS_VOLTAGE_DAC),
    ("U5",       1,  "Op-amp (bias V-to-I)",
        "Dual precision RRIO op-amp; closes high-side PMOS transconductance loop",
        "Vos < 0.5 mV, RRIO, 3.3 V", "MSOP-8 or smaller",
        "", PROMPT_OP_AMP),
    ("Q1, Q2",   2,  "PMOS (bias V-to-I)",
        "Small-signal logic-level P-channel MOSFET pass element; high-side current source into BIASx",
        "|Vgs(th)| <= 2 V, Vds >= 6 V", "SOT-23 or SOT-666",
        "", PROMPT_PMOS),
    ("Rsense1, Rsense2", 2, "Resistor 5.11 kOhm precision",
        "V-to-I sense resistor; sets I_FS = (3.3V - V_DAC) / R_sense",
        "5.11 kOhm, 0.1%, 25 ppm/C, thin-film", "0603",
        "", PROMPT_RSENSE),
    ("Qen1, Qen2 (DNP)", 2, "NMOS enable FET (DNP)",
        "Optional series enable FET between PMOS drain and BIASx jumper; hard isolation",
        "Vgs(th) <= 2 V, Vds >= 6 V", "SOT-23",
        "2N7002",
        "Optional / DNP. Series NMOS pass switch between Q1/Q2 PMOS drain and BIASx jumper. GPIO low at reset -> NMOS off -> hard output isolation independent of MCP4728 EEPROM state."),
    ("J1-J7",    7,  "RF connector",
        "Board-mount SMA jack, 50 ohm, DC-6 GHz",
        "SMA female, 50 ohm", "edge or vertical TH",
        "", PROMPT_SMA),
    ("J8",       1,  "FMC connector",
        "VITA 57.1 FMC LPC mezzanine-side connector",
        "160-pin LPC, 1.27 mm pitch", "FMC LPC, SMT",
        "", PROMPT_FMC_CONNECTOR),
    ("J9",       1,  "Pin header",
        "1x4 100-mil pin header for GPIO breakout",
        "1x4, 2.54 mm pitch", "TH vertical",
        "", PROMPT_HEADER_1X4),
    ("JP1-JPn",  8,  "Jumper header + shunt",
        "1x2 100-mil header with mating shunt cap",
        "1x2, 2.54 mm pitch", "TH vertical",
        "", PROMPT_JUMPER_1X2),
    ("C_bulk*",  4,  "Bulk MLCC",
        "10 uF X7R bulk decoupling, one per Bobcat supply pin bank",
        "10 uF, X7R, >=6.3 V", "0603 or 0805",
        "", PROMPT_DECOUPLE_CAP),
    ("C_mid*",   4,  "Mid MLCC",
        "1 uF X7R mid-band decoupling, one per Bobcat supply pin bank",
        "1 uF, X7R, >=6.3 V", "0402 or 0603",
        "", PROMPT_DECOUPLE_CAP),
    ("C_hf*",    4,  "HF MLCC",
        "100 nF X7R high-frequency decoupling, one per Bobcat supply pin bank",
        "100 nF, X7R, >=6.3 V", "0402",
        "", PROMPT_DECOUPLE_CAP),
    ("R0*",     20,  "Resistor 0 ohm",
        "Series 0 ohm jumpers on FMC LA signals and analog rails",
        "0 ohm", "0402",
        "", PROMPT_ZERO_OHM),
    ("R10k*",   12,  "Resistor 10 kOhm",
        "Pull-up / pull-down resistors on digital control lines",
        "10 kOhm, 5%", "0402",
        "", PROMPT_10K_PULL),
    ("TP1-TP4",  4,  "Test point",
        "GND test clips / loops",
        "GND probe point", "TH loop",
        "", PROMPT_TEST_POINT),
]


def build_workbook(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    for row_idx, (ref, qty, cat, desc, value, pkg, mpn, prompt) in enumerate(ROWS, start=2):
        row_data = [
            row_idx - 1,    # Item
            ref,            # Ref Des
            qty,            # Qty
            cat,            # Category
            desc,           # Description
            value,          # Value / Rating
            pkg,            # Package
            mpn,            # MPN
            "",             # Datasheet URL
            "",             # voltai chat url
            "",             # Notes
            prompt,         # Parts Agent Prompt
        ]
        for col_idx, value_ in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value_)
            cell.alignment = body_align
            cell.border = border
        ws.row_dimensions[row_idx].height = max(60, 12 * (prompt.count("\n") + 1)) if prompt else 30

    ws.freeze_panes = "B2"

    # --- Tab 2: per-refdes BOM (Part ref | Value | Description) ----------
    netlist_dir = Path(__file__).resolve().parent / "netlist"
    per_ref = _collect_per_refdes(netlist_dir)
    ws2 = wb.create_sheet("Per-Refdes")
    cols2 = [("Part ref", 12), ("Value", 16), ("MPN", 24),
             ("Description", 60), ("Sheet", 12)]
    for col_idx, (name, width) in enumerate(cols2, start=1):
        c = ws2.cell(row=1, column=col_idx, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[1].height = 22
    for row_idx, (ref, value, mpn, desc, sheet) in enumerate(per_ref, start=2):
        for col_idx, v in enumerate((ref, value, mpn, desc, sheet), start=1):
            c = ws2.cell(row=row_idx, column=col_idx, value=v)
            c.alignment = body_align
            c.border = border
    ws2.freeze_panes = "A2"

    wb.save(path)
    return len(ROWS), len(per_ref)


if __name__ == "__main__":
    out = str(Path(__file__).resolve().parent / "test1_bom.xlsx")
    n_main, n_ref = build_workbook(out)
    print(f"Wrote {out}  (BOM={n_main} rows, Per-Refdes={n_ref} rows)")
